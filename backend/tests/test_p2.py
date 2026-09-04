"""Регрессии P2: read_only парсер, кэш дат/PDF, interpret по запросу, start/stop."""
from pathlib import Path

import pandas as pd

from services import chat_service, data_tools, pdf_export
from services.excel_parser import parse_excel


REPO_ROOT = Path(__file__).resolve().parents[2]


class TestParserReadOnly:
    def test_load_workbook_uses_read_only(self, tmp_path, monkeypatch):
        from openpyxl import Workbook
        import services.excel_parser as ep

        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"
        ws.append(["клиент", "сумма"])
        ws.append(["Альфа", 100])
        path = tmp_path / "p2.xlsx"
        wb.save(path)

        seen = {}
        real = ep.openpyxl.load_workbook

        def wrapped(*args, **kwargs):
            seen.update(kwargs)
            return real(*args, **kwargs)

        monkeypatch.setattr(ep.openpyxl, "load_workbook", wrapped)
        sheets = parse_excel(str(path))
        assert seen.get("read_only") is True
        assert seen.get("data_only") is True
        assert seen.get("keep_links") is False
        assert "Данные" in sheets
        assert len(sheets["Данные"]) == 1


class TestDateColumnsCache:
    def test_second_call_skips_to_datetime(self, monkeypatch):
        df = pd.DataFrame(
            {
                "дата": ["01.01.2024", "02.01.2024", "03.01.2024"],
                "сумма": [10, 20, 30],
            }
        )
        data_tools._DATE_COLUMNS_CACHE.clear()
        calls = {"n": 0}
        real = data_tools.pd.to_datetime

        def wrapped(*args, **kwargs):
            calls["n"] += 1
            return real(*args, **kwargs)

        monkeypatch.setattr(data_tools.pd, "to_datetime", wrapped)
        first = data_tools.detect_date_columns(df)
        n_first = calls["n"]
        assert n_first >= 1
        assert "дата" in first["columns"]
        second = data_tools.detect_date_columns(df)
        assert calls["n"] == n_first
        assert second == first


class TestInterpretOnDemand:
    def test_stat_without_hint_skips_interpret(self, sales_df, monkeypatch):
        called = {"n": 0}

        def boom(*_a, **_k):
            called["n"] += 1
            return "лишняя интерпретация"

        monkeypatch.setattr(chat_service, "_maybe_interpret", boom)
        result = chat_service._execute_actions(
            sales_df,
            "Сколько строк?",
            [{"action": "stat", "operation": "row_count"}],
        )
        assert called["n"] == 0
        assert "лишняя интерпретация" not in result["answer"]
        assert "2394" in result["answer"]

    def test_stat_with_hint_appends_note(self, sales_df, monkeypatch):
        monkeypatch.setattr(
            chat_service, "_maybe_interpret", lambda *_a, **_k: "пояснение XYZ"
        )
        result = chat_service._execute_actions(
            sales_df,
            "Сколько строк? поясни",
            [{"action": "stat", "operation": "row_count"}],
        )
        assert "пояснение XYZ" in result["answer"]
        assert "2394" in result["answer"]

    def test_fast_path_does_not_interpret(self, sales_df, monkeypatch):
        monkeypatch.setattr(
            chat_service, "_maybe_interpret", lambda *_a, **_k: "не должно попасть"
        )
        result = chat_service.handle_question(sales_df, "Сколько строк в таблице?")
        assert "не должно попасть" not in result["answer"]
        assert "2394" in result["answer"]


class TestPdfCache:
    def test_same_key_computes_once(self, monkeypatch):
        from services import report_service

        df = pd.DataFrame({"сумма": [1, 2, 3]})
        pdf_export._pdf_cache.clear()
        counts = {"get": 0, "render": 0}

        def fake_get(*_a, **_k):
            counts["get"] += 1
            return {
                "report_type": "deficit_report",
                "metadata": {"filename": "x.xlsx", "rows": 3, "columns": 1},
                "narrative": "n",
                "kpis": [],
                "insights": [],
                "data_quality": {
                    "total_cells": 3,
                    "null_cells": 0,
                    "null_pct": 0.0,
                    "duplicates": 0,
                    "worst_columns": [],
                },
                "columns_overview": [],
                "charts": [],
            }

        def fake_render(*_a, **_k):
            counts["render"] += 1
            return b"%PDF-cached"

        monkeypatch.setattr(report_service, "get_full_report", fake_get)
        monkeypatch.setattr(pdf_export, "render_report_pdf", fake_render)
        monkeypatch.setattr(pdf_export, "load_dashboard_tabs", lambda *_a, **_k: [])

        first = pdf_export.cached_report_pdf(
            "file-a", df, filename="x.xlsx", narrative=None, insights=None, comment=None
        )
        second = pdf_export.cached_report_pdf(
            "file-a", df, filename="x.xlsx", narrative=None, insights=None, comment=None
        )
        assert first == second == b"%PDF-cached"
        assert counts["get"] == 1
        assert counts["render"] == 1

        pdf_export.cached_report_pdf(
            "file-a",
            df,
            filename="x.xlsx",
            narrative="другой текст",
            insights=None,
            comment=None,
        )
        assert counts["get"] == 2
        assert counts["render"] == 2


class TestStartStopScripts:
    def test_start_waits_for_backend_health(self):
        bat = (REPO_ROOT / "start.bat").read_text(encoding="ascii")
        assert "Waiting for backend" in bat
        assert "http://127.0.0.1:8000/" in bat
        ps1 = (REPO_ROOT / "start.ps1").read_text(encoding="ascii")
        assert "Waiting for backend" in ps1
        assert "http://127.0.0.1:8000/" in ps1

    def test_stop_only_kills_project_processes(self):
        text = (REPO_ROOT / "stop.ps1").read_text(encoding="ascii")
        assert "Test-ExcelAgentProcess" in text
        assert r"uvicorn\s+app:app" in text
        assert "streamlit" in text
        assert "Skip PID" in text
        assert "not Excel Agent" in text
        bat = (REPO_ROOT / "stop.bat").read_text(encoding="ascii")
        assert "stop.ps1" in bat
