"""Юнит-тесты 1C-aware парсера на реальных выгрузках."""
import pandas as pd

from services.data_tools import detect_date_columns
from services.excel_parser import parse_excel

from conftest import DEFICIT_FILE, SALES_FILE


class TestSalesFile:
    def test_shape(self, sales_df):
        assert sales_df.shape == (2394, 34)

    def test_key_columns_present(self, sales_df):
        for col in ("подразделение", "компания", "ответственный",
                    "дата начала сделки", "сумма по сделке", "статус"):
            assert col in sales_df.columns, f"нет колонки {col}"

    def test_no_metadata_rows(self, sales_df):
        # строки-заголовки 1С («Период», «Отбор» и т.п.) не должны попадать в данные
        first_col = sales_df.columns[0]
        values = sales_df[first_col].astype(str).str.lower()
        assert not values.str.contains("период:").any()
        assert not values.str.contains("отбор:").any()

    def test_date_column_detected(self, sales_df):
        detected = detect_date_columns(sales_df)["columns"]
        assert "дата начала сделки" in detected

    def test_revenue_is_numeric(self, sales_df):
        assert pd.api.types.is_numeric_dtype(sales_df["сумма по сделке"])

    def test_revenue_total(self, sales_df):
        total = sales_df["сумма по сделке"].sum()
        # эталон из e2e-прогона 26.08
        assert abs(total - 18_254_222_243.30) < 1.0


class TestDeficitFile:
    def test_shape(self, deficit_df):
        assert deficit_df.shape[0] > 0
        assert deficit_df.shape[1] >= 10

    def test_date_column_detected(self, deficit_df):
        detected = detect_date_columns(deficit_df)["columns"]
        assert detected, "в отчёте дефицита должна находиться колонка даты"

    def test_has_numeric_column(self, deficit_df):
        numeric = deficit_df.select_dtypes(include=["number"]).columns
        assert len(numeric) >= 1


class TestDeterminism:
    def test_parse_twice_same_shape(self):
        sheets1 = parse_excel(str(SALES_FILE))
        sheets2 = parse_excel(str(SALES_FILE))
        assert list(sheets1) == list(sheets2)
        for name in sheets1:
            assert sheets1[name].shape == sheets2[name].shape


class TestMultiSheetWorkbook:
    def test_skips_empty_and_hidden(self, tmp_path):
        from openpyxl import Workbook
        from services.excel_service import read_excel, read_workbook

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Сделки"
        ws1.append(["клиент", "сумма"])
        ws1.append(["Альфа", 100])
        ws2 = wb.create_sheet("Оплаты")
        ws2.append(["дата оплаты", "сумма оплаты"])
        ws2.append(["2024-01-15", 50])
        wb.create_sheet("Пустой")
        hidden = wb.create_sheet("Секрет")
        hidden.append(["скрытая колонка", "секрет"])
        hidden.append(["x", 1])
        hidden.sheet_state = "hidden"
        path = tmp_path / "multi.xlsx"
        wb.save(path)

        sheets = parse_excel(str(path))
        assert list(sheets) == ["Сделки", "Оплаты"]
        assert "сумма оплаты" in sheets["Оплаты"].columns
        assert "Пустой" not in sheets
        assert "Секрет" not in sheets

        df, workbook = read_workbook(str(path))
        assert list(workbook) == ["Сделки", "Оплаты"]
        assert list(df.columns) == list(workbook["Сделки"].columns)
        only_first = read_excel(str(path))
        assert "сумма оплаты" not in only_first.columns

