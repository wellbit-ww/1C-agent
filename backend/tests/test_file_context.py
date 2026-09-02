"""Карточка понимания файла: детерминированный брифинг, LLM, API, промпты."""
import json

import pandas as pd
import pytest
from fastapi.testclient import TestClient

from app import app
from models.file_context import FileContext
from services import chat_service, dashboard_service, db_service, file_context_service
from services.file_context_service import (
    data_hash,
    deterministic_context,
    ensure_context,
    enrich_with_llm,
)

from conftest import DEFICIT_FILE, SALES_FILE


class _FakeMessage:
    def __init__(self, content: str):
        self.content = content


class _FakeLLM:
    def __init__(self, content: str):
        self._content = content
        self.prompts: list[str] = []

    def invoke(self, prompt: str):
        self.prompts.append(prompt)
        return _FakeMessage(self._content)


@pytest.fixture
def tmp_db(tmp_path, monkeypatch):
    monkeypatch.setattr(db_service, "DB_PATH", tmp_path / "test.db")
    db_service.init_db()


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c


def _llm_payload(df: pd.DataFrame, **overrides) -> dict:
    data = {
        "title": "Тестовый отчёт",
        "summary": "Это карточка понимания файла для теста. Одна строка — запись выгрузки.",
        "grain": "одна строка = запись",
        "report_kind": "Тест",
        "metrics": list(df.select_dtypes(include="number").columns.astype(str)[:2]),
        "groupers": [
            str(c)
            for c in df.columns
            if not pd.api.types.is_numeric_dtype(df[c])
        ][:2],
        "date_columns": [],
        "caveats": ["тестовое ограничение"],
        "dashboard_ideas": ["Собери дашборд по клиентам"],
    }
    data.update(overrides)
    return data


class TestDeterministicContext:
    def test_metrics_are_real_columns(self, sales_df):
        ctx = deterministic_context(sales_df, filename="sales.xlsx")
        known = {str(c) for c in sales_df.columns}
        assert ctx.llm_ready is False
        assert ctx.summary
        assert set(ctx.metrics) <= known
        assert set(ctx.groupers) <= known
        assert ctx.metrics

    def test_csv_like_file(self):
        df = pd.DataFrame(
            {
                "клиент": ["А", "Б", "В"],
                "сумма": [100.0, 200.0, 50.0],
                "менеджер": ["Иванов", "Петров", "Иванов"],
            }
        )
        ctx = deterministic_context(df, filename="продажи.csv")
        known = set(df.columns)
        assert set(ctx.metrics) <= known
        assert set(ctx.groupers) <= known
        assert "сумма" in ctx.metrics
        assert ctx.llm_ready is False

    def test_hash_changes_with_columns(self, sales_df):
        other = sales_df.copy()
        other["новая"] = 1
        assert data_hash(sales_df) != data_hash(other)

    def test_persist_roundtrip(self, sales_df, tmp_db):
        ctx = ensure_context("f1", sales_df, filename="sales.xlsx", use_llm=False)
        loaded = file_context_service.get_context("f1")
        assert loaded is not None
        assert loaded.summary == ctx.summary
        assert loaded.llm_ready is False
        again = ensure_context("f1", sales_df, filename="sales.xlsx", use_llm=False)
        assert again.summary == ctx.summary


class TestLlmBriefing:
    def test_valid_json_becomes_llm_ready(self, sales_df, tmp_db, monkeypatch):
        payload = _llm_payload(sales_df)
        monkeypatch.setattr(
            file_context_service,
            "_get_brief_llm",
            lambda: _FakeLLM(json.dumps(payload, ensure_ascii=False)),
        )
        fallback = deterministic_context(sales_df, filename="sales.xlsx")
        ctx = enrich_with_llm(sales_df, fallback, filename="sales.xlsx")
        assert ctx.llm_ready is True
        assert "карточка понимания" in ctx.summary.lower()
        known = {str(c) for c in sales_df.columns}
        assert set(ctx.metrics) <= known

    def test_invented_columns_are_dropped(self, sales_df, monkeypatch):
        real = str(sales_df.select_dtypes(include="number").columns[0])
        payload = _llm_payload(
            sales_df,
            metrics=["выручка_xyz_несуществует", real],
            groupers=["менеджер_которого_нет"],
        )
        monkeypatch.setattr(
            file_context_service,
            "_get_brief_llm",
            lambda: _FakeLLM(json.dumps(payload, ensure_ascii=False)),
        )
        fallback = deterministic_context(sales_df)
        ctx = enrich_with_llm(sales_df, fallback)
        assert "выручка_xyz_несуществует" not in ctx.metrics
        assert real in ctx.metrics
        known = {str(c) for c in sales_df.columns}
        assert set(ctx.groupers) <= known

    def test_garbage_llm_falls_back(self, sales_df, monkeypatch):
        monkeypatch.setattr(
            file_context_service,
            "_get_brief_llm",
            lambda: _FakeLLM("я не понял файл <think>хм</think>"),
        )
        fallback = deterministic_context(sales_df, filename="sales.xlsx")
        ctx = enrich_with_llm(sales_df, fallback, filename="sales.xlsx")
        assert ctx.llm_ready is False
        assert ctx.summary == fallback.summary

    def test_ensure_saves_fallback_when_ollama_down(self, sales_df, tmp_db, monkeypatch):
        from services.exceptions import OllamaUnavailableError

        def boom(*a, **k):
            raise OllamaUnavailableError("нет ollama")

        monkeypatch.setattr(file_context_service, "enrich_with_llm", boom)
        ctx = ensure_context("f2", sales_df, filename="sales.xlsx", use_llm=True)
        assert ctx.llm_ready is False
        assert file_context_service.get_context("f2") is not None


class TestPromptWiring:
    def test_classify_prompt_includes_summary(self, sales_df, monkeypatch):
        captured = {}

        def fake_classify(prompt):
            captured["prompt"] = prompt
            return '{"action": "help"}'

        monkeypatch.setattr(chat_service, "classify", fake_classify)
        ctx = FileContext(
            summary="Секретный брифинг файла XYZ",
            metrics=[str(sales_df.columns[0])],
            dashboard_ideas=["Собери дашборд"],
        )
        chat_service.handle_question(
            sales_df,
            "что это за выгрузка на самом деле",
            file_context=ctx,
        )
        assert "Секретный брифинг файла XYZ" in captured["prompt"]

    def test_dashboard_prompt_includes_summary(self, sales_df, monkeypatch):
        captured = {}

        class CapturingLLM:
            def invoke(self, prompt):
                captured["prompt"] = prompt
                spec = {
                    "tabs": [{
                        "title": "Обзор",
                        "tiles": [{
                            "title": "Тест",
                            "chart_type": "bar",
                            "source": {
                                "kind": "group",
                                "group_column": str(sales_df.columns[0]),
                            },
                            "agg": "count",
                        }],
                    }]
                }
                return _FakeMessage(json.dumps(spec))

        monkeypatch.setattr(dashboard_service, "_get_spec_llm", CapturingLLM)
        ctx = FileContext(summary="Брифинг для дашборда ABC")
        dashboard_service.generate_spec_nl(
            sales_df, "собери дашборд по клиентам", file_context=ctx
        )
        assert "Брифинг для дашборда ABC" in captured["prompt"]

    def test_exact_column_chart(self, sales_df):
        from services.generic_dashboard import pick_groupers, pick_metrics

        groupers = pick_groupers(sales_df)
        metrics = pick_metrics(sales_df)
        assert groupers and metrics
        result = chat_service._exec_chart(
            sales_df,
            {
                "chart_type": "bar",
                "group_column": groupers[0],
                "value_column": metrics[0],
                "question": "график",
            },
        )
        assert result.get("chart")
        assert groupers[0] in result["answer"]


class TestApi:
    def test_upload_does_not_call_llm(self, client, monkeypatch):
        called = []

        def boom(*a, **k):
            called.append(1)
            raise AssertionError("upload must not wait for LLM")

        monkeypatch.setattr(file_context_service, "enrich_with_llm", boom)
        with open(DEFICIT_FILE, "rb") as f:
            response = client.post(
                "/upload", files={"file": ("deficit.xlsx", f)}
            )
        assert response.status_code == 200, response.text
        assert not called

    def test_dashboard_includes_deterministic_context(self, client):
        with open(SALES_FILE, "rb") as f:
            uploaded = client.post("/upload", files={"file": ("sales.xlsx", f)})
        assert uploaded.status_code == 200
        file_id = uploaded.json()["file_id"]
        dash = client.post("/dashboard", json={"file_id": file_id})
        assert dash.status_code == 200, dash.text
        ctx = dash.json().get("file_context") or {}
        assert ctx.get("summary")
        assert ctx.get("llm_ready") is False

    def test_file_context_endpoint_mocked_llm(self, client, monkeypatch):
        with open(SALES_FILE, "rb") as f:
            uploaded = client.post("/upload", files={"file": ("sales.xlsx", f)})
        file_id = uploaded.json()["file_id"]
        from services.excel_service import read_excel

        df = read_excel(str(SALES_FILE))
        payload = _llm_payload(df)
        monkeypatch.setattr(
            file_context_service,
            "_get_brief_llm",
            lambda: _FakeLLM(json.dumps(payload, ensure_ascii=False)),
        )
        response = client.post("/file-context", json={"file_id": file_id})
        assert response.status_code == 200, response.text
        body = response.json()
        assert body["llm_ready"] is True
        assert "карточка понимания" in body["summary"].lower()


def _multi_sheet_path(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Сделки"
    ws1.append(["клиент", "сумма"])
    ws1.append(["Альфа", 100])
    ws1.append(["Бета", 200])
    ws2 = wb.create_sheet("Оплаты")
    ws2.append(["дата оплаты", "сумма оплаты", "документ"])
    ws2.append(["2024-01-15", 50, "Платёжка-1"])
    ws2.append(["2024-02-01", 80, "Платёжка-2"])
    wb.create_sheet("Пустой")
    hidden = wb.create_sheet("Секрет")
    hidden.append(["скрытая колонка", "секретная сумма"])
    hidden.append(["не должно попасть в LLM", 999])
    hidden.sheet_state = "hidden"
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    return path


class TestLlmSeesAllSheets:
    """Снимок для модели должен содержать каждый видимый непустой лист."""

    def test_snapshot_lists_every_visible_sheet(self, tmp_path):
        from services.excel_parser import parse_excel
        from services.file_context_service import build_snapshot, catalog_sheets

        path = _multi_sheet_path(tmp_path)
        workbook = parse_excel(str(path))
        df = next(iter(workbook.values()))
        snap = build_snapshot(df, filename="multi.xlsx", workbook=workbook)

        names = [s["name"] for s in snap["sheets"]]
        assert names == ["Сделки", "Оплаты"]
        assert snap["active_sheet"] == "Сделки"
        assert snap["sheets"][0]["active"] is True
        assert snap["sheets"][1]["active"] is False
        assert "сумма оплаты" in snap["sheets"][1]["columns"]
        assert "документ" in snap["sheets"][1]["columns"]
        dumped = json.dumps(snap, ensure_ascii=False)
        assert "Оплаты" in dumped
        assert "сумма оплаты" in dumped
        assert "Секрет" not in dumped
        assert "скрытая колонка" not in dumped
        assert catalog_sheets(workbook)[1]["name"] == "Оплаты"

    def test_wide_first_sheet_does_not_drop_later_sheets(self, tmp_path):
        from openpyxl import Workbook
        from services.excel_parser import parse_excel
        from services.file_context_service import build_snapshot

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Широкий"
        ws1.append([f"колонка_{i}" for i in range(40)])
        ws1.append(list(range(40)))
        ws2 = wb.create_sheet("Второй")
        ws2.append(["уникальная_метрика_листа2", "группа"])
        ws2.append([123, "А"])
        path = tmp_path / "wide.xlsx"
        wb.save(path)
        workbook = parse_excel(str(path))
        df = next(iter(workbook.values()))
        snap = build_snapshot(df, filename="wide.xlsx", workbook=workbook)
        payload = json.dumps(snap, ensure_ascii=False, default=str)[:12000]
        assert payload.startswith("{")
        assert "Второй" in payload
        assert "уникальная_метрика_листа2" in payload

    def test_deterministic_context_names_all_sheets(self, tmp_path):
        from services.excel_parser import parse_excel
        from services.file_context_service import deterministic_context

        workbook = parse_excel(str(_multi_sheet_path(tmp_path)))
        df = next(iter(workbook.values()))
        ctx = deterministic_context(df, filename="multi.xlsx", workbook=workbook)
        assert [s.name for s in ctx.sheets] == ["Сделки", "Оплаты"]
        assert ctx.active_sheet == "Сделки"
        assert "Оплаты" in ctx.summary
        assert any("Оплаты" in c for c in ctx.caveats)
        block = ctx.prompt_block()
        assert "Оплаты" in block
        assert "сумма оплаты" in block

    def test_llm_prompt_contains_second_sheet(self, tmp_path, monkeypatch):
        from services.excel_parser import parse_excel
        from services.file_context_service import deterministic_context, enrich_with_llm

        workbook = parse_excel(str(_multi_sheet_path(tmp_path)))
        df = next(iter(workbook.values()))
        fallback = deterministic_context(df, filename="multi.xlsx", workbook=workbook)
        llm = _FakeLLM(json.dumps(_llm_payload(df), ensure_ascii=False))
        monkeypatch.setattr(file_context_service, "_get_brief_llm", lambda: llm)
        enrich_with_llm(df, fallback, filename="multi.xlsx", workbook=workbook)
        prompt = llm.prompts[0]
        assert "Оплаты" in prompt
        assert "сумма оплаты" in prompt
        assert "Сделки" in prompt
        assert "Секрет" not in prompt

    def test_llm_cannot_drop_parsed_sheets(self, tmp_path, monkeypatch):
        from services.excel_parser import parse_excel
        from services.file_context_service import deterministic_context, enrich_with_llm

        workbook = parse_excel(str(_multi_sheet_path(tmp_path)))
        df = next(iter(workbook.values()))
        fallback = deterministic_context(df, filename="multi.xlsx", workbook=workbook)
        payload = _llm_payload(df, summary="Только один лист сделок.")
        monkeypatch.setattr(
            file_context_service,
            "_get_brief_llm",
            lambda: _FakeLLM(json.dumps(payload, ensure_ascii=False)),
        )
        ctx = enrich_with_llm(df, fallback, filename="multi.xlsx", workbook=workbook)
        assert [s.name for s in ctx.sheets] == ["Сделки", "Оплаты"]
        assert any("Оплаты" in c for c in ctx.caveats)

    def test_fixture_files_all_parsed_sheets_reach_snapshot(self, sales_df):
        from services.excel_parser import parse_excel
        from services.file_context_service import build_snapshot

        workbook = parse_excel(str(SALES_FILE))
        snap = build_snapshot(sales_df, filename="sales.xlsx", workbook=workbook)
        assert {s["name"] for s in snap["sheets"]} == set(workbook)
        assert len(snap["sheets"]) == len(workbook)

        workbook_d = parse_excel(str(DEFICIT_FILE))
        from services.excel_service import read_excel

        deficit = read_excel(str(DEFICIT_FILE))
        snap_d = build_snapshot(deficit, filename="deficit.xlsx", workbook=workbook_d)
        assert {s["name"] for s in snap_d["sheets"]} == set(workbook_d)

    def test_upload_persists_all_sheets(self, client, tmp_path):
        path = _multi_sheet_path(tmp_path)
        with open(path, "rb") as f:
            response = client.post(
                "/upload", files={"file": ("multi.xlsx", f)}
            )
        assert response.status_code == 200, response.text
        file_id = response.json()["file_id"]
        dash = client.post("/dashboard", json={"file_id": file_id})
        assert dash.status_code == 200, dash.text
        ctx = dash.json().get("file_context") or {}
        names = [s["name"] for s in ctx.get("sheets") or []]
        assert names == ["Сделки", "Оплаты"]
        assert ctx.get("active_sheet") == "Сделки"
        assert "Оплаты" in (ctx.get("summary") or "")
